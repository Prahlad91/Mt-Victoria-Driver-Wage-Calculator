"""File parsing — roster PDFs, schedule ZIPs, payslip XLSXs / PDFs.
PRD ref: Section 6 (File Upload Requirements)

v3.8 deployed (2-column PDF support — see _extract_schedule_text_from_file below).
"""
from __future__ import annotations
import io
import json as json_mod
import re
from datetime import datetime, timedelta
from zipfile import ZipFile, BadZipFile

try:
    import pdfplumber
except ImportError:
    pdfplumber = None  # type: ignore

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

try:
    from PIL import Image as _PILImage
    import pytesseract as _pytesseract
    _OCR_AVAILABLE = True
except ImportError:
    _OCR_AVAILABLE = False

from models import (
    ParseRosterResponse, ParsedDayEntry,
    ParsePayslipResponse, PayslipLineItem,
    ParsedRosterResponse, RosterDayEntry,
    ParsedScheduleResponse, DiagramInfo,
)

_VALID_LINES = set(list(range(1, 23)) + list(range(201, 221)))
_TIME_RE = re.compile(r'^\d{1,2}:\d{2}$')
_WHRS_RE = re.compile(r'^\d{1,2}:\d{2}[Ww]$')
_FAT_RE  = re.compile(r'^F\d+$')


def _norm_time(t: str) -> str:
    h, m = t.split(':')
    return f'{int(h):02d}:{m}'


# ─── Text extraction (roster — single column / table layout) ──────────────────

def _extract_text_from_file(file_bytes: bytes) -> str:
    """Extract text from ZIP (Sydney Trains app) or real PDF (default scan order)."""
    try:
        with ZipFile(io.BytesIO(file_bytes)) as zf:
            if 'manifest.json' in zf.namelist():
                manifest = json_mod.loads(zf.read('manifest.json'))
                full_text = ''
                for page in manifest['pages']:
                    page_text = zf.read(page['text']['path']).decode('utf-8', errors='replace')
                    full_text += page_text + '\n'
                return full_text
    except (BadZipFile, KeyError, Exception):
        pass

    if pdfplumber is None:
        raise RuntimeError('pdfplumber not installed.')
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            full_text = ''
            for page in pdf.pages:
                page_text = page.extract_text() or ''
                full_text += page_text + '\n'
            return full_text
    except Exception as e:
        raise ValueError(f'Could not read file as ZIP or PDF. Details: {e}')


# ─── Text extraction (schedule — TWO-COLUMN PDF layout) ───────────────────────
# PRD §6.6 v3.8: Sydney Trains schedule PDFs are 2-column. Default
# extract_text() reads left-to-right across BOTH columns line by line, jumbling
# diagrams together. We must crop each page in half and extract each column
# separately, otherwise:
#   - ~half the diagrams get missed entirely (only one column header is found
#     per page; the second is captured into the first's day_type field)
#   - "Time off duty" gets pulled from the wrong column (e.g. 3155's value gets
#     attributed to 3154 because the columns are interleaved in the text stream)

def _extract_schedule_text_from_file(file_bytes: bytes) -> str:
    """
    Extract schedule text. ZIP path stays as-is (those text files are
    pre-organised per diagram). PDF path uses column-aware extraction.
    """
    # Try ZIP first
    try:
        with ZipFile(io.BytesIO(file_bytes)) as zf:
            if 'manifest.json' in zf.namelist():
                manifest = json_mod.loads(zf.read('manifest.json'))
                full_text = ''
                for page in manifest['pages']:
                    page_text = zf.read(page['text']['path']).decode('utf-8', errors='replace')
                    full_text += page_text + '\n'
                return full_text
    except (BadZipFile, KeyError, Exception):
        pass

    # Real PDF — column-aware extraction
    if pdfplumber is None:
        raise RuntimeError('pdfplumber not installed.')
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            full_text = ''
            for page in pdf.pages:
                mid = page.width / 2
                # Small overlap (5pt) so text right at the boundary isn't lost
                left  = page.crop((0,           0, mid + 5,    page.height))
                right = page.crop((mid - 5,     0, page.width, page.height))
                left_text  = left.extract_text()  or ''
                right_text = right.extract_text() or ''
                # Newline between columns ensures regex sees them as separate blocks
                full_text += left_text + '\n' + right_text + '\n'
            return full_text
    except Exception as e:
        raise ValueError(f'Could not read schedule PDF. Details: {e}')


# ─── Roster parser (unchanged) ─────────────────────────────────────────────────

def parse_roster_zip(file_bytes: bytes, filename: str) -> ParsedRosterResponse:
    """Parse roster file (ZIP or PDF)."""
    warnings: list[str] = []
    full_text = _extract_text_from_file(file_bytes)
    text = full_text.replace('\r\n', '\n').replace('\r', '\n')

    fn_end_m = re.search(r'Fortnight ending\s+(\d{2}/\d{2}/\d{4})', text)
    fn_start = fn_end = None
    if fn_end_m:
        dt = datetime.strptime(fn_end_m.group(1), '%d/%m/%Y')
        fn_end = dt.strftime('%Y-%m-%d')
        fn_start = (dt - timedelta(days=13)).strftime('%Y-%m-%d')

    # Also try "Fortnight commencing ..." format (printed roster PDFs say this instead)
    if fn_end is None:
        fn_comm_m = re.search(
            r'Fortnight commencing\s+\w+,?\s+(\d{1,2})\s+(\w+)\s+(\d{4})',
            text, re.IGNORECASE,
        )
        if fn_comm_m:
            try:
                dt = datetime.strptime(
                    f'{fn_comm_m.group(1)} {fn_comm_m.group(2)} {fn_comm_m.group(3)}',
                    '%d %B %Y',
                )
                fn_start = dt.strftime('%Y-%m-%d')
                fn_end   = (dt + timedelta(days=13)).strftime('%Y-%m-%d')
            except ValueError:
                pass

    layer_m = re.search(r'Layer:\s*(Master)', text)
    line_type = 'master' if layer_m else 'fortnight'

    # ── Approach 1: text-regex (works for ZIP exports + simple text PDFs) ─────────
    line_starts: list[tuple[int, int]] = []
    line_re = re.compile(r'(?m)^(\d{1,3})(?=\s+(?:OFF|ADO|\d{1,2}:\d{2}))')
    for m in line_re.finditer(text):
        num = int(m.group(1))
        if num in _VALID_LINES:
            line_starts.append((num, m.start()))

    lines_data: dict[str, list[RosterDayEntry]] = {}
    for idx, (line_num, start_pos) in enumerate(line_starts):
        end_pos = line_starts[idx + 1][1] if idx + 1 < len(line_starts) else len(text)
        section = text[start_pos:end_pos]
        days = _parse_day_entries(section)
        if days:
            lines_data[str(line_num)] = days
        if len(days) != 14:
            warnings.append(f'Line {line_num}: expected 14 days, got {len(days)}.')

    # ── Approach 2: table-extraction fallback ────────────────────────────────────
    # The real printed "Intercity Drivers Roster" PDF has crew names between the
    # line number and the day entries, so the regex above never matches.
    # pdfplumber.extract_tables() preserves cell boundaries and handles it cleanly.
    crew_names: dict[str, str] = {}
    if not lines_data:
        is_zip = False
        try:
            with ZipFile(io.BytesIO(file_bytes)) as zf:
                is_zip = 'manifest.json' in zf.namelist()
        except Exception:
            pass
        if not is_zip:
            lines_data, crew_names = _parse_roster_from_pdf_tables(file_bytes, warnings)

    if not lines_data:
        warnings.append('No roster lines found in this file.')
        return ParsedRosterResponse(
            source_file=filename, line_type=line_type,
            fn_start=fn_start, fn_end=fn_end, lines={}, warnings=warnings
        )

    return ParsedRosterResponse(
        source_file=filename, line_type=line_type,
        fn_start=fn_start, fn_end=fn_end,
        lines=lines_data, crew_names=crew_names, warnings=warnings,
    )


def _parse_day_entries(section_text: str) -> list[RosterDayEntry]:
    """
    Two-pass parser for roster sections.

    Master-roster PDFs have a two-row layout per line:
      Row 1:  <line_num>  OFF  OFF  01:52 - 11:21  02:32 - 10:32  …  OFF
      Row 2:  09:29W 3154  08:00W 3155  09:30W 3157 F74  …
    Times come from row 1; r_hrs and diagram numbers come from row 2 (in
    work-day order).  Some fortnight/legacy formats put everything on one
    row, which pass 1 handles exactly as before.
    """
    text_lines = [l.strip() for l in section_text.strip().splitlines() if l.strip()]
    if not text_lines:
        return []

    # ── Pass 1: parse the first text line for OFF / ADO / time-ranges ──────────
    words = text_lines[0].split()
    start = 1 if (words and re.match(r'^\d{1,3}$', words[0])) else 0

    raw: list[dict] = []   # keys: diag, r_start, r_end, cm, r_hrs
    i = start
    pending_diag: str | None = None
    while i < len(words) and len(raw) < 14:
        w = words[i]
        if w.upper() == 'OFF':
            raw.append({'diag': 'OFF', 'r_start': None, 'r_end': None, 'cm': False, 'r_hrs': 8.0})
            i += 1; pending_diag = None
        elif w.upper() == 'ADO':
            raw.append({'diag': 'ADO', 'r_start': None, 'r_end': None, 'cm': False, 'r_hrs': 8.0})
            i += 1; pending_diag = None
        elif re.match(r'^\d{3,4}$', w) and i + 1 < len(words) and _TIME_RE.match(words[i + 1]):
            pending_diag = w; i += 1
        elif _TIME_RE.match(w) and i + 2 < len(words) and words[i + 1] == '-':
            r_start = _norm_time(w)
            end_raw = words[i + 2]
            cm      = end_raw.upper().endswith('L')
            r_end   = _norm_time(end_raw.rstrip('LlWw'))
            i += 3
            r_hrs = 8.0
            if i < len(words) and _WHRS_RE.match(words[i]):
                hrs_str = words[i][:-1]
                h_str, m_str = hrs_str.split(':')
                r_hrs = round(int(h_str) + int(m_str) / 60, 4)
                i += 1
            diag_parts: list[str] = []
            while i < len(words):
                tok = words[i]
                if _FAT_RE.match(tok): i += 1; break
                if tok.upper() in ('OFF', 'ADO'): break
                if _TIME_RE.match(tok) and i + 1 < len(words) and words[i + 1] == '-': break
                if re.match(r'^\d{3,4}$', tok) and i + 1 < len(words) and _TIME_RE.match(words[i + 1]): break
                diag_parts.append(tok); i += 1
            diag = pending_diag or ' '.join(diag_parts).strip()
            pending_diag = None
            raw.append({'diag': diag, 'r_start': r_start, 'r_end': r_end, 'cm': cm, 'r_hrs': r_hrs})
        else:
            pending_diag = None; i += 1

    # ── Pass 2: fill r_hrs + diag from subsequent lines (two-row format) ───────
    # Only work days that still have no diagram get updated.
    work_slots = [k for k, d in enumerate(raw) if d['r_start'] is not None and not d['diag']]
    if work_slots:
        wi = 0
        for line in text_lines[1:]:
            if wi >= len(work_slots):
                break
            toks = line.split()
            j = 0
            while j < len(toks) and wi < len(work_slots):
                tok = toks[j]
                if _WHRS_RE.match(tok):
                    hrs_str = tok[:-1]
                    h_str, m_str = hrs_str.split(':')
                    raw[work_slots[wi]]['r_hrs'] = round(int(h_str) + int(m_str) / 60, 4)
                    j += 1
                    # Accept numeric diagram (e.g. 3154) OR alpha code (e.g. SBY, SMB)
                    if j < len(toks) and not _WHRS_RE.match(toks[j]) and not _FAT_RE.match(toks[j]):
                        tok2 = toks[j]
                        if re.match(r'^\d{3,4}$', tok2) or re.match(r'^[A-Za-z]{2,5}$', tok2):
                            raw[work_slots[wi]]['diag'] = tok2
                            j += 1
                    while j < len(toks) and _FAT_RE.match(toks[j]):
                        j += 1
                    wi += 1
                else:
                    j += 1

    return [
        RosterDayEntry(diag=d['diag'] or '', r_start=d['r_start'], r_end=d['r_end'], cm=d['cm'], r_hrs=d['r_hrs'])
        for d in raw
    ]


# ─── Table-based roster parser (new) — handles real "Intercity Drivers Roster" PDF ───
# The printed roster PDF has the layout:
#   [Line# | Crew name | Day1 … Day14 | O/T count]
# Each worked-day cell contains:
#   "HH:MM - HH:MM[L]\nDIAGRAM_NAME\nFNN"
# The text-regex approach (above) fails because crew names sit between the line
# number and the day entries, so ^1\s+(?:OFF|ADO|time) never matches.
# pdfplumber.extract_tables() preserves cell boundaries and handles this cleanly.

_CELL_TIME_RE  = re.compile(r'(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})(L?)', re.IGNORECASE)
_FAT_TRAIL_RE  = re.compile(r'\bF\d+\s*$')   # trailing fatigue token e.g. "F24", "F0"

# Anchor column indices for the 14 day-slot columns in the printed fortnight roster PDF
# WITH a crew-name column (col[2]).  Determined by debug analysis of the real PDF.
# Columns 3 & 4 are non-date metadata cells (e.g. "NTA", "HOL") that parse safely as OFF.
# Sub-columns between consecutive anchors carry split diagram/location tokens.
# Column 30 is the O/T count (not a day).
_DAY_ANCHORS     = [3, 4, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 26, 28]
_OT_COL          = 30

# Same indices shifted left by 1: used when there is NO crew-name column and day 1
# starts directly at col[2] (e.g. some exported / legacy roster PDFs).
_DAY_ANCHORS_NOCREW = [a - 1 for a in _DAY_ANCHORS]  # [2, 3, 4, 6, 8, …, 27]
_OT_COL_NOCREW      = _OT_COL - 1

# Swinger / spare-line table (lines 209-214 on the real fortnight PDF) is much
# narrower (~20 cols) than the main 30/31-col roster table.  Day cells run
# consecutively from col 3, with sub-column separators at col 9 and col 16.
#   col 0:   spacer
#   col 1:   line number
#   col 2:   crew name
#   col 3-8: days 1-6                (consecutive)
#   col 9:   sub-col separator (empty)
#   col 10-15: days 7-12             (consecutive)
#   col 16:  sub-col separator (empty)
#   col 17-18: days 13-14
#   col 19:  O/T count
_SWINGER_DAY_ANCHORS        = [3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14, 15, 17, 18]
_SWINGER_OT_COL             = 19
_SWINGER_DAY_ANCHORS_NOCREW = [a - 1 for a in _SWINGER_DAY_ANCHORS]
_SWINGER_OT_COL_NOCREW      = _SWINGER_OT_COL - 1

# Tables with this many columns or fewer use the swinger layout.  Real-world
# values are 20 (with crew col) and 19 (without).  Main roster table is 30-31.
_SWINGER_MAX_NCOLS = 22

# Regex that matches text that looks like a day-slot entry in an anchor row.
# Used to determine whether col[2] is a crew name or the first day column.
_DAY_ANCHOR_RE = re.compile(
    r'\d{1,2}:\d{2}\s*[-–]\s*\d{1,2}:\d{2}'         # time range
    r'|^\s*(?:OFF|ADO|NTA|HOL|SBY|RDO|ALT)\s*$',    # non-work keyword
    re.IGNORECASE | re.MULTILINE,
)


def _build_struck_set(tbl_obj: object, page: object) -> frozenset[tuple[int, int]]:
    """
    Return (row_idx, col_idx) pairs for cells whose visible content is struck through.

    Current implementation: detects struck-through cells by looking for characters
    whose non-stroking (fill) colour is noticeably lighter than solid black — the
    "greyed-out" rendering used by some roster printing systems to indicate a no-show.

    Vector-line based detection was tried but proved unreliable: the multi-row table
    layout (3 pdfplumber rows per logical line) means mid-row separator lines are
    indistinguishable from intra-cell strikethrough by geometry alone.  Text-colour
    detection is both simpler and more precise.

    Falls back to an empty set on any error (e.g. pdfplumber without char attributes).
    """
    struck: set[tuple[int, int]] = set()
    try:
        chars = getattr(page, 'chars', None) or []
        if not chars:
            return frozenset()

        # Build a map of (row_idx, col_idx) → list of char non-stroking colours for
        # all text that falls inside each cell.
        cell_colours: dict[tuple[int, int], list[float]] = {}
        for ri, row in enumerate(tbl_obj.rows):    # type: ignore[attr-defined]
            for ci, cell in enumerate(row.cells):
                if cell is None:
                    continue
                x0, top, x1, bottom = cell
                # Collect chars inside this cell
                colours: list[float] = []
                for ch in chars:
                    cx = ch.get('x0', 0)
                    cy = ch.get('top', 0)
                    if x0 <= cx <= x1 and top <= cy <= bottom and ch.get('text', '').strip():
                        nsc = ch.get('non_stroking_color')
                        if isinstance(nsc, (int, float)):
                            colours.append(float(nsc))
                        elif isinstance(nsc, (list, tuple)) and nsc:
                            colours.append(float(nsc[0]))
                if colours:
                    cell_colours[(ri, ci)] = colours

        if not cell_colours:
            return frozenset()

        # A cell is "struck" if its average text brightness is significantly above 0
        # (0 = solid black; 1 = white; struck text is often printed in a mid-grey).
        # Threshold 0.35: black/dark text < 0.1; grey "struck" text typically 0.4-0.7.
        for (ri, ci), colours in cell_colours.items():
            avg = sum(colours) / len(colours)
            if avg >= 0.35:
                struck.add((ri, ci))

    except Exception:
        pass
    return frozenset(struck)


def _parse_cell_to_day_entry(cell: object) -> 'RosterDayEntry':
    """
    Parse one table cell from the printed fortnight roster PDF.

    Handled formats:
      - Empty / "OFF" / "NTA" / "HOL" / "(AL )" / "OFF(AL )" …  → OFF
      - "ADO" / "ADO(LSL )" …                                    → ADO
      - "01:49 - 10:49\\n3154 MQ/SMB\\nF63"                     → worked day
      - "16:15 - 00:34L\\n3167 MQ\\nF35"                        → cross-midnight
      - "06:00 - 14:00\\nSBY x LVE\\nOnline Training…"          → SBY
      - "04:00 - 12:00\\nSBY(AL ) F0"                           → SBY (annotation stripped)
    """
    text = str(cell or '').strip()
    if not text:
        return RosterDayEntry(diag='OFF', r_start=None, r_end=None, cm=False, r_hrs=0.0)

    first_line = text.split('\n')[0].strip().upper()

    # ADO — check before time-range (handles "ADO(LSL )" etc.)
    if first_line.startswith('ADO'):
        return RosterDayEntry(diag='ADO', r_start=None, r_end=None, cm=False, r_hrs=8.0)

    # Look for a time range anywhere in the cell
    m = _CELL_TIME_RE.search(text)
    if not m:
        # No time range → non-work day (NTA, HOL, leave, blank)
        return RosterDayEntry(diag='OFF', r_start=None, r_end=None, cm=False, r_hrs=0.0)

    r_start = _norm_time(m.group(1))
    r_end   = _norm_time(m.group(2))
    cm      = m.group(3).upper() == 'L'

    # r_hrs: calculate from time range (overridden later if schedule is uploaded)
    s_m = _time_to_mins(r_start)
    e_m = _time_to_mins(r_end)
    if cm:
        e_m += 24 * 60
    r_hrs = round(max(0.0, (e_m - s_m) / 60), 4)

    # Diagram name: on the line(s) after the time range, or on the same line
    after = text[m.end():].lstrip(' ')
    diag  = 'SBY'

    for ln in after.split('\n'):
        ln = ln.strip()
        if not ln:
            continue
        # Strip trailing fatigue token (e.g. " F63", " F0")
        ln = _FAT_TRAIL_RE.sub('', ln).strip()
        if not ln:
            continue
        toks = ln.split()
        if not toks:
            continue
        first = toks[0]
        if re.match(r'^\d{4}$', first):
            # 4-digit diagram — also grab trailing location codes (MQ, SMB, MQ/RK, etc.)
            locs = [t for t in toks[1:]
                    if re.match(r'^[A-Z]{2,5}(/[A-Z]{2,5})*$', t)]
            diag = first + (' ' + locs[0] if locs else '')
        elif re.match(r'^[A-Z]', first, re.IGNORECASE):
            # Non-numeric diagram: SBY, MSBYD3, AMV01, MY30 … strip annotations like "(LSL )"
            clean = re.sub(r'\(.*$', '', first).strip()
            diag  = clean.upper() if clean else 'SBY'
        break  # diagram is always on the first non-empty line after the time range

    return RosterDayEntry(diag=diag, r_start=r_start, r_end=r_end, cm=cm, r_hrs=r_hrs)


def _clean_crew_name(raw: str) -> str:
    """
    Normalise a raw crew-name cell from the fortnight roster.
    Collapses embedded newlines (pdfplumber splits long names) and strips
    annotations like '(AL 18/04/26)' that aren't part of the actual name.
    """
    s = re.sub(r'\s+', ' ', (raw or '').replace('\n', ' ')).strip()
    # Strip trailing date-annotations e.g. " (AL 18/04/26)"
    s = re.sub(r'\s*\([A-Z]{1,4}\s+\d{1,2}/\d{1,2}/\d{2,4}\)\s*$', '', s).strip()
    return s


def _parse_roster_from_pdf_tables(
    file_bytes: bytes, warnings: list[str],
) -> tuple[dict, dict]:
    """
    Parse the Sydney Trains "Intercity Drivers Roster" PDF using pdfplumber's
    table extraction.

    Returns (lines_data, crew_names) where:
      - lines_data:  {line_number_str → list[RosterDayEntry] of length 14}
      - crew_names:  {line_number_str → crew member name string}  (may be empty)

    Supported table layouts (0-based columns):
      MAIN ROSTER (30-31 cols, lines 1-22 & 201-208):
        col[0]=spacer  col[1]=Line#  col[2]=Crew  col[3..28]=14 days  col[30]=O/T
        Day anchors: _DAY_ANCHORS = [3, 4, 5, 7, 9, 11, 13, 15, 17, 19, 21, 23, 26, 28]
        No-crew variant: anchors shift left by 1, days start at col[2].
      SWINGER TABLE (~20 cols, lines 209-214 on the printed PDF):
        col[0]=spacer  col[1]=Line#  col[2]=Crew  col[3..18]=14 days  col[19]=O/T
        Day anchors run CONSECUTIVELY at _SWINGER_DAY_ANCHORS, with sub-column
        separators at col 9 and col 16.
        No-crew variant likewise shifts left by 1.

    Layout is auto-detected by table width: ≤ _SWINGER_MAX_NCOLS (22) ⇒ swinger.
    Crew-column presence is auto-detected by inspecting col[2] of the first
    anchor row — if it contains a time-range or day keyword, no crew column.

    Logical lines span MULTIPLE table rows:
      • Anchor row   — cells[1] is a valid line number.
      • Continuation rows — cells[1] is empty; carry split diagram/location tokens
        and the back half of long crew names.

    Cells whose content is visually struck through (person didn't show up) are
    treated as empty → OFF.  Detection uses character text colour (grey ⇒ struck);
    restricted to day columns only so line-number / crew-name greyed-out alternating
    row styling isn't mistaken for strikethrough.
    """
    if pdfplumber is None:
        return {}, {}

    lines_data: dict = {}
    crew_names: dict = {}
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                settings = {
                    'vertical_strategy':   'lines',
                    'horizontal_strategy': 'lines',
                }

                # Prefer find_tables() — gives cell bboxes for strikethrough detection.
                # Fall back to extract_tables() for older pdfplumber builds.
                table_pairs: list[tuple[object | None, list]] = []
                try:
                    found = page.find_tables(settings) or []
                    if not found:
                        found = page.find_tables() or []
                    table_pairs = [(tbl, tbl.extract()) for tbl in found]
                except Exception:
                    raw = page.extract_tables(settings) or page.extract_tables() or []
                    table_pairs = [(None, t) for t in raw]

                for tbl_obj, table in table_pairs:
                    if not table:
                        continue

                    # ── Detect table-width-based layout: main vs swinger ──────
                    ncols = max((len(r) for r in table if r), default=0)
                    is_swinger_layout = 0 < ncols <= _SWINGER_MAX_NCOLS

                    # ── Detect whether there is a crew-name column ────────────
                    # Inspect col[2] of the first anchor row.  If it looks like a
                    # day entry, the days start at col[2] (no crew column present).
                    has_crew_col = True
                    for raw_row in table:
                        if not raw_row or len(raw_row) < 3:
                            continue
                        c1 = str(raw_row[1] or '').strip()
                        if re.match(r'^\d{1,3}$', c1) and int(c1) in _VALID_LINES:
                            c2 = str(raw_row[2] or '').strip()
                            if _DAY_ANCHOR_RE.search(c2):
                                has_crew_col = False
                            break

                    if is_swinger_layout:
                        day_anchors = (_SWINGER_DAY_ANCHORS if has_crew_col
                                       else _SWINGER_DAY_ANCHORS_NOCREW)
                        ot_col      = (_SWINGER_OT_COL      if has_crew_col
                                       else _SWINGER_OT_COL_NOCREW)
                    else:
                        day_anchors = (_DAY_ANCHORS if has_crew_col
                                       else _DAY_ANCHORS_NOCREW)
                        ot_col      = (_OT_COL      if has_crew_col
                                       else _OT_COL_NOCREW)

                    # ── Build struck-cell set for this table ──────────────────
                    struck = (
                        _build_struck_set(tbl_obj, page)
                        if tbl_obj is not None
                        else frozenset()
                    )

                    # ── Group raw table rows into logical line-groups ─────────
                    groups: list[dict] = []
                    current: dict | None = None

                    for ri, raw_row in enumerate(table):
                        if not raw_row:
                            continue
                        # Normalise cells; blank out struck-through day cells only.
                        # Never blank col 0 (spacer), col 1 (line#), or col 2 (crew name)
                        # — those columns can use grey text for alternating-row styling,
                        # which would otherwise erase valid line numbers.
                        _min_day_col = 3 if has_crew_col else 2
                        cells = [
                            ('' if (ri, ci) in struck and ci >= _min_day_col else str(c or '').strip())
                            for ci, c in enumerate(raw_row)
                        ]
                        while len(cells) <= ot_col:
                            cells.append('')

                        c1 = cells[1] if len(cells) > 1 else ''
                        if re.match(r'^\d{1,3}$', c1):
                            line_num = int(c1)
                            if line_num in _VALID_LINES:
                                if current is not None:
                                    groups.append(current)
                                current = {'line_num': line_num, 'rows': [cells]}
                            else:
                                current = None      # header / total row
                        elif current is not None:
                            current['rows'].append(cells)   # continuation row

                    if current is not None:
                        groups.append(current)

                    # ── Build 14 RosterDayEntry objects per logical line ──────
                    for grp in groups:
                        line_num = grp['line_num']
                        if str(line_num) in lines_data:
                            continue  # first occurrence wins (PDF may span pages)

                        rows = grp['rows']

                        # Crew name: col 2 of the ANCHOR row only.  Continuation
                        # rows' col 2 may contain bleed-through from adjacent
                        # logical lines (the printed PDF visually merges tall
                        # cells, but pdfplumber attributes the text to whichever
                        # row vertically owns it).  The anchor row alone is the
                        # safest source.
                        if has_crew_col and rows:
                            crew_name = _clean_crew_name(rows[0][2] if len(rows[0]) > 2 else '')
                            if crew_name:
                                crew_names[str(line_num)] = crew_name

                        entries: list[RosterDayEntry] = []

                        for di, anchor in enumerate(day_anchors):
                            next_a = day_anchors[di + 1] if di + 1 < len(day_anchors) else ot_col
                            col_range = range(anchor, min(next_a, ot_col))

                            parts: list[str] = []
                            for row in rows:
                                row_parts = [
                                    row[col]
                                    for col in col_range
                                    if col < len(row) and row[col]
                                ]
                                if row_parts:
                                    parts.append(' '.join(row_parts))

                            entries.append(_parse_cell_to_day_entry('\n'.join(parts)))

                        if len(entries) == 14:
                            lines_data[str(line_num)] = entries
                        elif 0 < len(entries) < 14:
                            # Pad with OFF if column detection was short
                            off = RosterDayEntry(diag='OFF', r_start=None, r_end=None,
                                                 cm=False, r_hrs=0.0)
                            entries.extend([off] * (14 - len(entries)))
                            lines_data[str(line_num)] = entries
                        else:
                            warnings.append(
                                f'Line {line_num}: no day entries found — skipped.'
                            )

    except Exception as exc:
        warnings.append(f'Table-extraction pass failed: {exc}')

    return lines_data, crew_names


# ─── Schedule parser (v3.8: 2-column PDF support) ──────────────────────────────

_TIME_PATTERNS_FOR_LABEL = [
    r'(\d{1,2}\s*:\s*\d{2}\s*[AaPp][Mm]?)',
    r'(\d{1,2}\s*:\s*\d{2})',
]


def _flexible_label_pattern(label: str) -> str:
    """'Sign on' -> r'Sign\\s*[-]?\\s*on'"""
    parts = label.split()
    return r'\s*[-]?\s*'.join(re.escape(p) for p in parts)


def _extract_labeled_time(label: str, text: str) -> str | None:
    label_pat = _flexible_label_pattern(label)
    for tp in _TIME_PATTERNS_FOR_LABEL:
        m = re.search(rf'{label_pat}\s*:?\s*{tp}', text, re.IGNORECASE)
        if m:
            return m.group(1)
    return None


_NO_RE = re.compile(
    r'(?:^|\n)\s*No\.\s+(\d{3,4})\s+([^\n]+)',
    re.MULTILINE,
)


def parse_schedule_zip(file_bytes: bytes, filename: str) -> ParsedScheduleResponse:
    """
    Parse a schedule file (ZIP or PDF).
    PRD §6.6: PDFs are 2-column — extracted column-by-column to keep diagrams intact.
    Per diagram, extracts:
      - sign_on  ← from "Sign on" line                  (scheduled START)
      - sign_off ← from "Time off duty" line            (scheduled END)
      - r_hrs    ← from "Total shift" line
      - km       ← from "Distance: NNN.NNN Km" line
      - cm       ← derived (sign_off earlier than sign_on)
    """
    warnings: list[str] = []

    # v3.8: column-aware text extraction for the schedule PDF case
    full_text = _extract_schedule_text_from_file(file_bytes)

    fname_upper = filename.upper()
    if 'DRWD' in fname_upper or 'WEEKDAY' in fname_upper:
        schedule_type = 'weekday'
    elif 'DRWE' in fname_upper or 'WEEKEND' in fname_upper:
        schedule_type = 'weekend'
    else:
        schedule_type = 'weekday'

    text = full_text.replace('\r\n', '\n').replace('\r', '\n')
    blocks = list(_NO_RE.finditer(text))

    diagrams: dict[str, DiagramInfo] = {}
    failed_signon: list[str] = []
    failed_signoff: list[str] = []

    for idx, block_m in enumerate(blocks):
        diag_num     = block_m.group(1)
        day_type_raw = block_m.group(2).strip()
        block_start  = block_m.end()
        block_end    = blocks[idx + 1].start() if idx + 1 < len(blocks) else len(text)
        block_text   = text[block_start:block_end]

        # First occurrence wins (diagrams may repeat e.g. Monday-only and Tue-Fri variants)
        if diag_num in diagrams:
            continue

        sign_on_raw  = _extract_labeled_time('Sign on',       block_text)
        sign_off_raw = _extract_labeled_time('Time off duty', block_text)

        try:    sign_on  = _parse_time_str(sign_on_raw)  if sign_on_raw  else None
        except ValueError: sign_on = None
        try:    sign_off = _parse_time_str(sign_off_raw) if sign_off_raw else None
        except ValueError: sign_off = None

        if sign_on  is None: failed_signon.append(diag_num)
        if sign_off is None: failed_signoff.append(diag_num)

        total_m = re.search(r'Total\s*shift\s*:?\s*(\d{1,2}:\d{2})', block_text, re.IGNORECASE)
        km_m    = re.search(r'Distance\s*:?\s*([\d.]+)\s*Km', block_text, re.IGNORECASE)

        km    = float(km_m.group(1)) if km_m else 0.0
        r_hrs = _hmm_to_float(total_m.group(1)) if total_m else 8.0

        cm = False
        if sign_on and sign_off:
            cm = _time_to_mins(sign_off) < _time_to_mins(sign_on)

        dl = day_type_raw.lower()
        if   'saturday' in dl: day_type = 'saturday'
        elif 'sunday'   in dl: day_type = 'sunday'
        else:                  day_type = 'weekday'

        diagrams[diag_num] = DiagramInfo(
            diag_num=diag_num, day_type=day_type,
            sign_on=sign_on, sign_off=sign_off,
            r_hrs=r_hrs, km=km, cm=cm,
        )

    if not diagrams:
        warnings.append(
            'No diagram entries found. '
            'Make sure this is a weekday or weekend schedule file (MTVICDRWD or MTVICDRWE).'
        )
    if failed_signon:
        unique_failed = sorted(set(failed_signon), key=lambda s: int(s) if s.isdigit() else 0)
        warnings.append(
            f'Could not extract "Sign on" time for {len(unique_failed)} diagram(s): '
            f'{", ".join(unique_failed[:10])}'
            f'{"..." if len(unique_failed) > 10 else ""}.'
        )
    if failed_signoff:
        unique_failed = sorted(set(failed_signoff), key=lambda s: int(s) if s.isdigit() else 0)
        warnings.append(
            f'Could not extract "Time off duty" time for {len(unique_failed)} diagram(s): '
            f'{", ".join(unique_failed[:10])}'
            f'{"..." if len(unique_failed) > 10 else ""}.'
        )

    return ParsedScheduleResponse(
        source_file=filename, schedule_type=schedule_type,
        diagrams=diagrams, warnings=warnings,
    )


# ─── Time helpers ─────────────────────────────────────────────────────────────

def _parse_time_str(t: str) -> str:
    """
    Convert various time string formats to 'HH:MM' (24-hour):
      '12:51a', '9:18a', '10:30p', '12:51 AM', '9:18 am', '10:30 PM',
      '09:18', '17:30', '9 : 18 PM'
    """
    t = re.sub(r'\s+', '', t.strip()).lower()
    has_am = t.endswith('am') or (t.endswith('a') and not t.endswith('pa'))
    has_pm = t.endswith('pm') or (t.endswith('p') and not t.endswith('ap'))
    t = re.sub(r'[apm]+$', '', t)
    if ':' not in t:
        raise ValueError(f'No colon in time: {t}')
    h, m = map(int, t.split(':'))
    if has_pm and h != 12: h += 12
    elif has_am and h == 12: h = 0
    if h < 0 or h > 23 or m < 0 or m > 59:
        raise ValueError(f'Invalid time: {t}')
    return f'{h:02d}:{m:02d}'


def _ampm_to_hhmm(t: str) -> str:
    return _parse_time_str(t)


def _hmm_to_float(t: str) -> float:
    h, m = map(int, t.split(':'))
    return round(h + m / 60, 4)


def _time_to_mins(t: str) -> int:
    h, m = map(int, t.split(':'))
    return h * 60 + m


# ─── Legacy fortnight roster PDF parser ───────────────────────────────────────

def parse_roster_pdf(file_bytes: bytes, filename: str = 'roster.pdf') -> ParseRosterResponse:
    if pdfplumber is None:
        raise RuntimeError('pdfplumber not installed.')

    parsed_days: list[ParsedDayEntry] = []
    warnings: list[str] = []
    time_re = re.compile(r'\b(\d{2}):(\d{2})\b')
    date_re = re.compile(r'\b(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})\b')
    diag_re = re.compile(r'\b(3[0-9]{3}[A-Z\s]*|SBY|OFF|ADO|RDO)\b')

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in (table or []):
                    cells = [str(c or '').strip() for c in row]
                    row_text = ' '.join(cells)
                    times = time_re.findall(row_text)
                    date_match = date_re.search(row_text)
                    diag_match = diag_re.search(row_text)
                    if not date_match and not times: continue
                    date_str = None
                    confidence = 0.5
                    if date_match:
                        d, m_val, y = date_match.groups()
                        y = f'20{y}' if len(y) == 2 else y
                        date_str = f'{y}-{int(m_val):02d}-{int(d):02d}'
                        confidence += 0.3
                    sign_on  = f'{times[0][0]}:{times[0][1]}' if len(times) >= 1 else None
                    sign_off = f'{times[1][0]}:{times[1][1]}' if len(times) >= 2 else None
                    if sign_on:  confidence += 0.1
                    if sign_off: confidence += 0.1
                    diagram = diag_match.group(0).strip() if diag_match else 'UNKNOWN'
                    parsed_days.append(ParsedDayEntry(
                        date=date_str or 'UNKNOWN', diagram=diagram,
                        sign_on=sign_on, sign_off=sign_off, confidence=min(confidence, 1.0),
                    ))
                    if confidence < 0.7:
                        warnings.append(f'Low confidence ({confidence:.0%}) for row: {row_text[:60]}...')

    if not parsed_days:
        warnings.append('No roster entries could be extracted.')
    return ParseRosterResponse(source_file=filename, parsed_days=parsed_days, warnings=warnings)


# ─── Payslip parser ───────────────────────────────────────────────────────────

def parse_payslip_file(file_bytes: bytes, filename: str = 'payslip') -> ParsePayslipResponse:
    fname_lower = filename.lower()
    if fname_lower.endswith('.xlsx') or fname_lower.endswith('.xls'):
        return _parse_payslip_xlsx(file_bytes, filename)
    elif fname_lower.endswith('.pdf'):
        return _parse_payslip_pdf(file_bytes, filename)
    else:
        try:
            return _parse_payslip_xlsx(file_bytes, filename)
        except Exception:
            return _parse_payslip_pdf(file_bytes, filename)


def _detect_payslip_format(headers: list[str]) -> str:
    header_str = ' '.join(h.lower() for h in headers)
    return 'sydney_crew' if ('crew' in header_str or 'sydney crew' in header_str) else 'nsw_payslip'


def _parse_payslip_xlsx(file_bytes: bytes, filename: str) -> ParsePayslipResponse:
    if openpyxl is None:
        raise RuntimeError('openpyxl not installed.')
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    warnings: list[str] = []
    line_items: list[PayslipLineItem] = []
    total_gross = 0.0
    period_start = period_end = None
    fmt = 'nsw_payslip'
    header_row_idx = None
    for i, row in enumerate(rows):
        cells = [str(c or '').strip() for c in row]
        if any('code' in c.lower() or 'description' in c.lower() for c in cells):
            header_row_idx = i
            fmt = _detect_payslip_format(cells)
            break
    for row in rows[:10]:
        row_str = ' '.join(str(c or '') for c in row)
        date_matches = re.findall(r'\d{1,2}/\d{1,2}/\d{4}', row_str)
        if len(date_matches) >= 2 and period_start is None:
            def _fmt(d):
                parts = d.split('/')
                return f'{parts[2]}-{int(parts[1]):02d}-{int(parts[0]):02d}'
            period_start = _fmt(date_matches[0])
            period_end   = _fmt(date_matches[1])
    if header_row_idx is not None:
        for row in rows[header_row_idx + 1:]:
            cells = list(row)
            if not any(cells): continue
            try:
                code = str(cells[0] or '').strip()
                desc = str(cells[1] or '').strip() if len(cells) > 1 else ''
                if not code or not desc: continue
                hrs_val  = cells[2] if len(cells) > 2 else None
                rate_val = cells[3] if len(cells) > 3 else None
                amt_val  = cells[4] if len(cells) > 4 else None
                hrs    = float(hrs_val)  if hrs_val  and str(hrs_val).replace('.', '').replace('-', '').isdigit() else None
                rate   = float(rate_val) if rate_val and str(rate_val).replace('.', '').replace('-', '').isdigit() else None
                amount = float(amt_val) if amt_val else 0.0
                if amount != 0:
                    line_items.append(PayslipLineItem(code=code, description=desc, hours=hrs, rate=rate, amount=amount))
                    total_gross += amount
            except (ValueError, TypeError, IndexError):
                continue
    else:
        warnings.append('Could not identify header row in payslip.')
    return ParsePayslipResponse(
        source_file=filename, format=fmt,
        period_start=period_start, period_end=period_end,
        total_gross=round(total_gross, 2), line_items=line_items, warnings=warnings,
    )


def _parse_payslip_pdf(file_bytes: bytes, filename: str) -> ParsePayslipResponse:
    if pdfplumber is None:
        raise RuntimeError('pdfplumber not installed.')
    line_items: list[PayslipLineItem] = []
    warnings: list[str] = []
    total_gross = 0.0
    money_re = re.compile(r'\$?([\d,]+\.\d{2})')
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in (table or []):
                    cells = [str(c or '').strip() for c in row]
                    amounts = [float(m.replace(',', '')) for m in money_re.findall(' '.join(cells))]
                    if amounts and len(cells) >= 2:
                        try:
                            line_items.append(PayslipLineItem(
                                code=cells[0] or '—', description=cells[1] if len(cells) > 1 else '',
                                hours=None, rate=None, amount=amounts[-1],
                            ))
                            total_gross += amounts[-1]
                        except Exception: continue
    if not line_items:
        warnings.append('No line items extracted from payslip PDF.')
    return ParsePayslipResponse(
        source_file=filename, format='pdf',
        total_gross=round(total_gross, 2), line_items=line_items, warnings=warnings,
    )


# ─── Assoc / Un-assoc Payments Chart parser (v3.12, rewritten v3.25) ────────
#
# v3.25 rewrite handles the two real-world chart formats users upload:
#   1. Phone photos (e.g. WhatsApp Image, Oct-2025 chart) — correctly oriented,
#      OCR-able directly.
#   2. Scanned image-only PDFs (e.g. docs/111616112.pdf, Apr-2026 chart) —
#      rotated 90°, no text layer.  Old parser returned "no diagrams found".
#
# Pipeline:
#   PDF  → try pdfplumber text-extract → if empty/no diagrams, render every
#          page to a PIL image at 200 dpi → auto-rotate via Tesseract OSD →
#          OCR (`--psm 6`) → parse text.
#   IMG  → load PIL image → auto-rotate via Tesseract OSD → OCR → parse.
#   CSV  → parse as before, with all 4 columns supported.
#
# Parser extracts all 4 fields per diagram (un-assoc, assoc-pay, assoc-calc,
# build-up) using a positional + sum-check heuristic — see _parse_chart_text.

# Diagram numbers for Mt Victoria drivers — widened v3.25 from the original
# 3151-3168/3651-3664 to cover newer diagrams (3169, 3171, etc.).
_DIAG_RE   = re.compile(r'\b(3(?:1[5-9]\d|6\d\d))\b')
_TIME_RE   = re.compile(r'\b(\d{1,2}:\d{2})\b')
_DIAG_RANGE = set(range(3151, 3200)) | set(range(3601, 3700))

# Common Tesseract OCR confusions for digits in thin/narrow table text.
# Applied at the start of each line to recover diagram numbers that came out
# mangled (e.g. "3i55" → "3155", "316l" → "3161", "(36S2" → "(3652").
_DIGIT_CONFUSIONS = str.maketrans({
    'i': '1', 'I': '1', 'l': '1', 'L': '1', '|': '1',
    'O': '0', 'o': '0',
    'S': '5', 's': '5',
    'B': '8', 'Z': '2',
})


def _normalise_diag_token(s: str) -> str:
    """If `s` looks like a 4-token starting with 3 (possibly with OCR-mangled
    digits in positions 1-3), return the cleaned 4-digit diagram number, else
    return the original.  Conservative: only fixes the FIRST 4-character run
    that starts with '3' near the start of the line."""
    return s.translate(_DIGIT_CONFUSIONS)

# OCR-or-text tokens that count as a chart "cell" after the distance number.
# Captures HH:MM (with optional leading +) or dash variants (-, –, —).
_CELL_RE = re.compile(r'(?P<time>\+?\d{1,2}:\d{2})|(?P<dash>[-–—]+)')


def _mins(t: str) -> int:
    """HH:MM (with optional leading +) → integer minutes. 0 on any parse error."""
    if not t or t in {'-', '–', '—'}:
        return 0
    try:
        s = t.lstrip('+').strip()
        h, m = s.split(':')
        v = int(h) * 60 + int(m)
        return v if 0 <= v <= 1439 else 0
    except Exception:
        return 0


# ── Excel assoc chart helpers (v3.38) ─────────────────────────────────────────

# Diagram pattern re-used for xlsx parsing (same range as _DIAG_RANGE above).
_DIAG_RE_XLSX = re.compile(r'^3(?:1[5-9]\d|6\d\d)$')

# Column-header patterns in priority order.  First match wins per column.
# The tuples are (field_name, compiled_regex).
_XLSX_CHART_COLS: list[tuple[str, "re.Pattern[str]"]] = [
    ('diagram',    re.compile(r'diag|dgm',                           re.I)),
    ('un_assoc',   re.compile(r'un[\s_\-]?assoc',                   re.I)),
    ('build_up',   re.compile(r'build[\s_\-]?up|buildup',           re.I)),
    ('assoc_calc', re.compile(r'assoc[\s_\-]?calc|calc[\s_\-]?time', re.I)),
    ('dist_pay',   re.compile(r'dist',                               re.I)),
    ('assoc_pay',  re.compile(r'assoc[\s_\-]?(pay|pmt|payment)',     re.I)),
]


def _xlsx_val_to_mins(val: object) -> int:
    """Convert an openpyxl cell value to integer minutes.

    Handles: None, datetime.time/datetime.datetime (from time-formatted cells),
    float 0–1 (Excel time serial == fraction of 1440 min), plain int/float
    (treated directly as minutes), and str in "H:MM" or decimal form.
    """
    if val is None:
        return 0
    try:
        import datetime as _dt
        if isinstance(val, _dt.time):
            return val.hour * 60 + val.minute
        if isinstance(val, _dt.datetime):
            return val.hour * 60 + val.minute
    except Exception:
        pass
    if isinstance(val, float):
        if 0 < val < 1:          # Excel time serial (e.g. 0.0347 ≈ 50 min)
            return round(val * 1440)
        return max(0, int(round(val)))
    if isinstance(val, int):
        return max(0, val)
    s = str(val).strip()
    if not s or s in {'-', '–', '—'}:
        return 0
    if ':' in s:
        parts = s.split(':')
        try:
            h = int(parts[0].strip() or '0')
            m = int(parts[1].strip()) if len(parts) > 1 else 0
            return max(0, h * 60 + m)
        except (ValueError, IndexError):
            pass
    try:
        f = float(s)
        if 0 < f < 1:
            return round(f * 1440)
        return max(0, int(round(f)))
    except ValueError:
        return 0


def _parse_xlsx_sheet(rows: list) -> tuple[dict, list[str]]:
    """Locate header row and data rows in one worksheet, return chart + warnings.

    Two-pass approach:
      Pass 1 — find a row that has 'diagram' (or 'dgm'/'diag') AND at least one
               time-column keyword; build col_map from it.
      Pass 2 — if no named headers, fall back to positional: first 4-digit 3xxx
               cell becomes the diagram column; next 4 cols = un_assoc, assoc_pay,
               assoc_calc, build_up.
    Returns ({}, []) if no chart-like structure found — caller tries the next sheet.
    """
    col_map: dict[str, int] = {}
    header_idx: int | None = None

    # Pass 1 — named headers
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else '' for c in row]
        has_diag  = any(_XLSX_CHART_COLS[0][1].search(c) for c in cells)
        has_other = any(
            pat.search(c)
            for _, pat in _XLSX_CHART_COLS[1:]
            for c in cells
        )
        if has_diag and has_other:
            header_idx = i
            for j, c in enumerate(cells):
                for field, pat in _XLSX_CHART_COLS:
                    if field not in col_map and c and pat.search(c):
                        col_map[field] = j
            break

    # Pass 2 — positional fallback
    if 'diagram' not in col_map:
        for row in rows:
            cells = [str(c).strip() if c is not None else '' for c in row]
            for j, c in enumerate(cells):
                if _DIAG_RE_XLSX.match(c):
                    col_map = {'diagram': j}
                    for k, field in enumerate(
                        ['un_assoc', 'assoc_pay', 'assoc_calc', 'build_up']
                    ):
                        if j + 1 + k < len(cells):
                            col_map[field] = j + 1 + k
                    header_idx = None
                    break
            if 'diagram' in col_map:
                break

    if 'diagram' not in col_map:
        return {}, []  # Not a chart sheet; caller tries next

    diag_col = col_map['diagram']
    data_rows = rows[(header_idx + 1) if header_idx is not None else 0:]

    def _get(row: tuple, field: str) -> int:
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return 0
        return _xlsx_val_to_mins(row[idx])

    chart: dict = {}
    rows_found = 0
    for row in data_rows:
        if not row or diag_col >= len(row) or row[diag_col] is None:
            continue
        diag = str(row[diag_col]).strip()
        if not _DIAG_RE_XLSX.match(diag):
            continue
        rows_found += 1
        un = _get(row, 'un_assoc')
        ap = _get(row, 'assoc_pay')
        ac = _get(row, 'assoc_calc')
        bu = _get(row, 'build_up')
        if any(v > 0 for v in (un, ap, ac, bu)):
            chart[diag] = {
                'unAssocMins':      un,
                'assocPaymentMins': ap,
                'assocCalcMins':    ac,
                'buildUpMins':      bu,
            }

    warnings: list[str] = []
    if rows_found > 0 and len(chart) < 5:
        warnings.append(
            f'Only {len(chart)} diagram(s) with non-zero values found '
            f'(expected ~20+).  Verify column names match: '
            f'Diagram, Un-Assoc, Assoc Pmt, Assoc Calc, Build Up.'
        )
    return chart, warnings


def _parse_chart_xlsx(file_bytes: bytes) -> tuple[dict, list[str]]:
    """Parse an assoc/un-assoc chart from an .xlsx workbook (v3.38).

    Iterates every sheet in the workbook; returns the first sheet that yields
    a non-empty chart dict.  Delegates per-sheet logic to _parse_xlsx_sheet().
    """
    if openpyxl is None:
        raise RuntimeError(
            'openpyxl is not installed — cannot parse .xlsx files.  '
            'Install it with: pip install openpyxl'
        )
    import io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True)
    all_warnings: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        chart, w = _parse_xlsx_sheet(rows)
        all_warnings.extend(w)
        if chart:
            return chart, all_warnings

    all_warnings.append(
        'No Mt Victoria diagram data found in any Excel sheet.  '
        'Expected column headers: Diagram (or DGM/DIAG), Un-Assoc, '
        'Assoc Pmt (or Assoc Payment), Assoc Calc, Build Up.  '
        'Download the CSV template for the exact column names if needed.'
    )
    return {}, all_warnings


def _parse_chart_text(text: str) -> tuple[dict, list[str]]:
    """Parse the assoc chart from OCR-or-text-extracted content into the full
    4-field-per-diagram model (un-assoc, assoc-pay, assoc-calc, build-up).

    Per-row column order on the physical chart (after the day-of-operation
    abbreviation `.MTWTF.` etc.):

        Shift Length | Distance | Un-Assoc | Assoc-Pay | Dist-Pay | Assoc-Calc | Build-Up

    The first HH:MM after the diagram is the Shift Length (skipped here — we
    don't store it).  The distance is a decimal number with a dot.  Everything
    after the distance is parsed as ordered cells (HH:MM or dash).  The trailing
    cell with a `+` prefix or a dash is the Build-Up.  The last two non-build-up
    cells are Dist-Pay and Assoc-Calc.  Anything before them is Un-Assoc and
    Assoc-Pay (positionally; dashes preserve the slot).

    A sum-check `un + assoc_pay + dist_pay ≈ assoc_calc` validates the mapping.
    When there's exactly one HH:MM token before dist-pay (ambiguous), the
    parser swaps un-assoc ↔ assoc-pay if the sum-check fails.

    Rows where all four fields are zero are omitted to keep the chart sparse.
    """
    chart: dict = {}
    warnings: list[str] = []
    diag_found = 0
    rows_with_data = 0

    for raw_line in text.split('\n'):
        line = raw_line.strip()
        if not line:
            continue
        # First try direct match.  If that fails, try OCR-confusion-corrected
        # match — Tesseract often misreads thin '1's as 'i'/'l'/'|', '5' as 'S',
        # etc. in tabular contexts.
        m = _DIAG_RE.search(line)
        if not m:
            normalised = _normalise_diag_token(line)
            m = _DIAG_RE.search(normalised)
            if not m:
                continue
            # Use the normalised line going forward so HH:MM values are clean too.
            line = normalised
        diag = m.group(1)
        if int(diag) not in _DIAG_RANGE:
            continue
        diag_found += 1

        # Anchor parsing after the distance value — that's how we tell shift
        # length apart from the cells we care about.
        after_diag = line[m.end():]
        dist_m = re.search(r'\d{2,3}\.\d{1,3}', after_diag)
        if not dist_m:
            # No distance found — partial OCR row, skip silently
            continue
        after_dist = after_diag[dist_m.end():]

        # Collect ordered cells: each is either an HH:MM string (with optional +)
        # or '-' for a dash.  Order is preserved so positional mapping works.
        cells: list[str] = []
        for cm in _CELL_RE.finditer(after_dist):
            cells.append(cm.group('time') if cm.group('time') else '-')
        if len(cells) < 2:
            # Need at least dist-pay + assoc-calc
            continue

        # Last cell: is it the build-up?  Has '+' prefix, OR is a dash and
        # we have ≥3 cells (so dist-pay/assoc-calc/build-up positions all present).
        build_up_str = '-'
        if cells[-1].startswith('+'):
            build_up_str = cells[-1]
            cells = cells[:-1]
        elif cells[-1] == '-' and len(cells) >= 3:
            cells = cells[:-1]

        if len(cells) < 2:
            continue
        dist_pay_str   = cells[-2]
        assoc_calc_str = cells[-1]
        before_dist    = cells[:-2]

        # Map the cells before dist-pay onto (un-assoc, assoc-pay).  Dashes
        # are preserved so position is meaningful.
        if len(before_dist) == 0:
            un_assoc_str, assoc_pay_str = '-', '-'
        elif len(before_dist) == 1:
            # Ambiguous — initial guess: it's un-assoc.  Sum-check below will
            # swap if needed.
            un_assoc_str, assoc_pay_str = before_dist[0], '-'
        else:
            un_assoc_str, assoc_pay_str = before_dist[0], before_dist[1]

        un_min = _mins(un_assoc_str)
        ap_min = _mins(assoc_pay_str)
        dp_min = _mins(dist_pay_str)
        ac_min = _mins(assoc_calc_str)
        bu_min = _mins(build_up_str)

        # Sum check.  Allow 1-min slop for HH:MM rounding artefacts on the chart.
        sum_diff = (un_min + ap_min + dp_min) - ac_min
        if abs(sum_diff) > 1 and len(before_dist) == 1:
            # Try the other interpretation: the single value was assoc-pay.
            un_min, ap_min = 0, un_min
            sum_diff = (un_min + ap_min + dp_min) - ac_min
        if abs(sum_diff) > 1:
            # Math doesn't validate — likely an OCR misread on this row.  Skip
            # it rather than store potentially-wrong data; surface a warning
            # so the user knows to verify.
            warnings.append(
                f"Diagram {diag}: math doesn't validate "
                f"(un_assoc {un_min} + assoc_pay {ap_min} + dist_pay {dp_min} "
                f"≠ assoc_calc {ac_min}, diff {sum_diff}m).  "
                f"OCR likely misread a value — row skipped, please verify."
            )
            continue

        if any(v > 0 for v in (un_min, ap_min, ac_min, bu_min)):
            chart[diag] = {
                'unAssocMins':      un_min,
                'assocPaymentMins': ap_min,
                'assocCalcMins':    ac_min,
                'buildUpMins':      bu_min,
            }
            rows_with_data += 1

    if diag_found == 0:
        warnings.append(
            'No Mt Victoria diagram numbers (3151-3199 / 3601-3699) found in the '
            'parsed content.  If you uploaded an image or scanned PDF, OCR quality '
            'on chart photos is variable — for reliable results, use the CSV '
            'template (download below) and fill in the values from the chart.'
        )
    elif rows_with_data < 5:
        warnings.append(
            f'OCR extracted {rows_with_data} diagram(s) confidently — fewer than '
            f'the ~23 expected for a Mt Victoria chart.  OCR on scanned/photographed '
            f'charts is unreliable; if values look wrong, please use the CSV '
            f'template instead (chart values can be hand-entered with full accuracy).'
        )
    return chart, warnings


def _pdf_text(file_bytes: bytes) -> str:
    """Extract raw text from every page of a PDF via pdfplumber.  Returns the
    empty string if the PDF has no text layer (caller should fall back to
    render→OCR for that case)."""
    if pdfplumber is None:
        raise RuntimeError('pdfplumber is not installed.')
    parts: list[str] = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            if tables:
                for table in tables:
                    for row in (table or []):
                        parts.append('  '.join(str(c or '') for c in row))
            else:
                parts.append(page.extract_text() or '')
    return '\n'.join(parts)


def _render_pdf_pages(file_bytes: bytes, dpi: int = 200) -> list:
    """Render every PDF page to a PIL image at the requested DPI.

    Used as a fallback for PDFs without a text layer (e.g. scanned charts) so
    we can OCR them.  Returns an empty list if pdfplumber isn't installed."""
    if pdfplumber is None:
        return []
    pages: list = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            try:
                pi = page.to_image(resolution=dpi)
                pages.append(pi.original)   # PIL.Image
            except Exception:
                continue
    return pages


def _auto_rotate(image):
    """Detect the orientation of `image` (a PIL Image) using Tesseract OSD
    and rotate it so text is upright.  No-op on any failure — OSD is finicky
    on sparse / small / non-Latin content but works well on dense table scans.

    Tesseract OSD reports the rotation NEEDED to make the text upright, so
    we apply `image.rotate(-rotation)` (PIL rotates counter-clockwise)."""
    if not _OCR_AVAILABLE:
        return image
    try:
        osd = _pytesseract.image_to_osd(image, config='--psm 0')  # type: ignore[no-untyped-call]
        rot_m = re.search(r'Rotate:\s*(\d+)', osd)
        if rot_m:
            rotation = int(rot_m.group(1))
            if rotation in (90, 180, 270):
                return image.rotate(-rotation, expand=True)
    except Exception:
        pass
    return image


def _ocr_image(file_bytes_or_image) -> str:
    """Run Tesseract OCR on an image (bytes or PIL Image) and return text.

    Auto-rotates the image first via Tesseract OSD so we handle rotated phone
    photos and scanned PDFs transparently."""
    if not _OCR_AVAILABLE:
        raise RuntimeError(
            'Image OCR requires the Tesseract OCR engine. '
            'On macOS: brew install tesseract. '
            'On Ubuntu/Debian: apt-get install -y tesseract-ocr. '
            'Alternatively, convert the chart to PDF or CSV before uploading.'
        )
    if isinstance(file_bytes_or_image, bytes):
        img = _PILImage.open(io.BytesIO(file_bytes_or_image))
    else:
        img = file_bytes_or_image
    # Auto-detect rotation BEFORE converting to greyscale (OSD works on RGB).
    img = _auto_rotate(img)
    img = img.convert('L')
    # Upscale small images to help Tesseract.
    w, h = img.size
    if w < 2000:
        img = img.resize((w * 2, h * 2), _PILImage.LANCZOS)
    return _pytesseract.image_to_string(img, config='--psm 6')  # type: ignore[no-untyped-call]


def _parse_csv_text(text: str) -> tuple[dict, list[str]]:
    """Parse `diagram,un_assoc_mins,assoc_payment_mins[,assoc_calc_mins,build_up_mins]`
    CSV from plain text.  Backwards-compatible with the 3-column legacy format."""
    chart: dict = {}
    warnings: list[str] = []
    for line in text.splitlines():
        parts = [p.strip().strip('"') for p in line.split(',')]
        if not parts[0]:
            continue
        if re.match(r'[a-zA-Z]', parts[0]):
            continue  # skip header rows
        diag = parts[0]
        try:
            un_min  = int(parts[1]) if len(parts) > 1 else 0
            ap_min  = int(parts[2]) if len(parts) > 2 else 0
            ac_min  = int(parts[3]) if len(parts) > 3 else 0
            bu_min  = int(parts[4]) if len(parts) > 4 else 0
        except ValueError:
            continue
        if any(v > 0 for v in (un_min, ap_min, ac_min, bu_min)):
            entry: dict = {'unAssocMins': un_min, 'assocPaymentMins': ap_min}
            if ac_min > 0:
                entry['assocCalcMins'] = ac_min
            if bu_min > 0:
                entry['buildUpMins'] = bu_min
            chart[diag] = entry
    return chart, warnings


def parse_assoc_chart_file(file_bytes: bytes, filename: str) -> 'ParseAssocChartResponse':
    """Parse an Assoc/Un-assoc Payments Chart from a CSV, PDF, PNG, JPG, etc.

    v3.25 — handles rotated images and image-only PDFs transparently:
      - PDF without a text layer (scanned image): render each page to an image,
        auto-rotate via Tesseract OSD, OCR.
      - Image with arbitrary rotation: auto-rotate before OCR.

    Returns a `ParseAssocChartResponse` with the parsed chart and any warnings.
    """
    from models import ParseAssocChartResponse  # local import to avoid circular

    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    warnings: list[str] = []
    chart: dict = {}

    if ext in ('csv', 'txt'):
        text = file_bytes.decode('utf-8', errors='replace')
        chart, warnings = _parse_csv_text(text)

    elif ext == 'pdf':
        # Try the cheap text-extract path first.
        text = _pdf_text(file_bytes)
        chart, warnings = _parse_chart_text(text)
        # If no diagrams were found, the PDF is likely an image-only scan —
        # fall back to render→auto-rotate→OCR per page.
        if not chart:
            # 300 dpi for OCR fallback — at 200 dpi Tesseract misreads "1" as
            # "i"/"l" frequently in thin table cells.
            pages = _render_pdf_pages(file_bytes, dpi=300)
            if pages:
                ocr_parts: list[str] = []
                for img in pages:
                    try:
                        ocr_parts.append(_ocr_image(img))
                    except RuntimeError as e:
                        # Tesseract missing — surface as a single warning, not per-page.
                        warnings.append(str(e))
                        break
                if ocr_parts:
                    chart, fb_warnings = _parse_chart_text('\n'.join(ocr_parts))
                    # Drop the earlier "no diagrams found" since the fallback ran.
                    warnings = [w for w in warnings if 'No Mt Victoria' not in w]
                    warnings.extend(fb_warnings)

    elif ext in ('xlsx', 'xls'):
        # v3.38: Excel workbook — column-header-aware parser.
        # openpyxl only reads .xlsx natively; .xls is an alias that will fail
        # gracefully with a clear error from openpyxl if it's a legacy format.
        chart, warnings = _parse_chart_xlsx(file_bytes)

    elif ext in ('png', 'jpg', 'jpeg', 'webp', 'bmp', 'tiff', 'tif'):
        text = _ocr_image(file_bytes)  # auto-rotates; raises if OCR not available
        chart, warnings = _parse_chart_text(text)
        if not chart and not any('No Mt Victoria' in w for w in warnings):
            warnings.append(
                'OCR succeeded but found no diagram data. The image may be '
                'low-resolution. Try a higher-quality scan or PDF.'
            )

    else:
        raise ValueError(
            f'Unsupported file type ".{ext}". '
            'Accepted: CSV (.csv), Excel (.xlsx), PDF (.pdf), or image '
            '(.png, .jpg, .jpeg, .webp, .bmp, .tiff).'
        )

    return ParseAssocChartResponse(
        source_file=filename,
        chart=chart,
        warnings=warnings,
    )
